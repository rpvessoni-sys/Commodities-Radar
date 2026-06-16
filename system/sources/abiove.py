"""Coletor ABIOVE — Associacao Brasileira das Industrias de Oleos Vegetais.

Padrao de URL identificado:
- https://abiove.org.br/abiove_content/Abiove/YYYY.MM-producao_entrega.xlsx
  → balanco mensal soja/farelo/oleo (sheet 'projecoes_mensais')
- https://abiove.org.br/abiove_content/Abiove/YYYY.MM-materia_prima.xlsx
  → producao de biodiesel por materia prima (oleo soja, gordura, etc)
- https://abiove.org.br/abiove_content/Abiove/exp_YYYYMM.xlsx
  → exportacoes BR detalhadas por destino

Estrategia: ler pagina /estatisticas/ para descobrir URLs mais recentes
(estavel ao longo dos anos), baixar Excel, parsear via openpyxl.
"""
import re
from datetime import date
from pathlib import Path

from .base import BaseCollector
import config


COMMODITIES_SLUG = {
    "soja": "soja",
    "grao": "soja",
    "graos": "soja",
    "farelo": "farelo",
    "oleo": "oleo",
    "óleo": "oleo",
}

METRICAS_TARGET = {
    "estoque inicial": "estoque_inicial",
    "estoque final": "estoque_final",
    "processamento": "esmagamento",
    "esmagamento": "esmagamento",
    "exportação": "exportacao",
    "exportacao": "exportacao",
    "compras locais": "compras_locais",
    "aquisição": "aquisicao",
    "aquisicao": "aquisicao",
    "vendas país": "vendas_internas",
    "produção": "producao",
    "producao": "producao",
}


class AbioveCollector(BaseCollector):
    source_name = "abiove"
    cadence = "monthly"
    description = "ABIOVE — balanco BR soja/farelo/oleo mensal (projecoes + realizado)"
    enabled = True

    def fetch(self):
        try:
            import requests
            from bs4 import BeautifulSoup
            from openpyxl import load_workbook
        except ImportError as e:
            raise RuntimeError(f"dependencia faltando: {e}")

        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "pt-BR,pt;q=0.9",
        }

        # 1) Descobrir URL do producao_entrega mais recente
        try:
            r = requests.get("https://abiove.org.br/estatisticas/", headers=headers, timeout=30)
            r.raise_for_status()
            soup = BeautifulSoup(r.text, "lxml")
        except Exception as e:
            return [self._erro(f"index: {type(e).__name__}: {e}")]

        url_producao = None
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if "producao_entrega" in href and href.endswith(".xlsx"):
                if href.startswith("/"):
                    href = "https://abiove.org.br" + href
                if not href.startswith("http"):
                    href = "https://" + href.lstrip("htps:/").lstrip("/")
                url_producao = href
                break

        if not url_producao:
            return [self._erro("Link producao_entrega.xlsx nao encontrado na pagina")]

        # 2) Baixar
        try:
            r = requests.get(url_producao, headers=headers, timeout=60)
            r.raise_for_status()
            xlsx_bytes = r.content
        except Exception as e:
            return [self._erro(f"download xlsx: {type(e).__name__}: {e}")]

        # Extrair YYYY.MM do filename para data_referencia
        m = re.search(r"(\d{4})\.(\d{2})", url_producao)
        if m:
            ano_ref = m.group(1)
            mes_ref = m.group(2)
            data_ref = f"{ano_ref}-{mes_ref}-01"
        else:
            data_ref = date.today().isoformat()

        # Salvar copia
        out_path = config.DATA_DIR / f"abiove_producao_entrega_{data_ref}.xlsx"
        out_path.write_bytes(xlsx_bytes)

        # 3) Parsear sheet projecoes_mensais
        try:
            wb = load_workbook(filename=out_path, data_only=True)
        except Exception as e:
            return [self._erro(f"abrir xlsx: {type(e).__name__}: {e}")]

        # Priorizar match exato; fallback para sheets que comecam com 'proje'
        sheet_name = None
        for nome in wb.sheetnames:
            if nome.lower() == "projecoes_mensais":
                sheet_name = nome
                break
        if not sheet_name:
            for nome in wb.sheetnames:
                # 'projeções' (com acento) e variantes
                if nome.lower().startswith("proje") and "projetado" not in nome.lower():
                    sheet_name = nome
                    break

        if not sheet_name:
            return [self._erro(f"Sheet 'projecoes_mensais' nao encontrada. Sheets: {wb.sheetnames}")]

        ws = wb[sheet_name]
        results = self._parse_balanco(ws, data_ref)

        # Metadado
        results.append({
            "data_referencia": data_ref,
            "tipo": "metadata",
            "commodity": "abiove",
            "metrica": "release",
            "valor": 1,
            "unidade": "bool",
            "contexto": f"ABIOVE balanco {data_ref} | {url_producao}",
            "raw": {"url": url_producao, "sheet": sheet_name},
        })

        return results

    def _parse_balanco(self, ws, data_ref: str) -> list[dict]:
        """Le sheet projecoes_mensais.

        Estrutura observada:
        - L4: header com meses (Fev, Mar, Abr, Mai...)
        - L6: "1. Soja"
        - L7-L14: linhas de soja (estoque inicial, processamento, etc)
        - L16+: "2. Farelo"
        - L23+: "3. Oleo"
        """
        results = []
        ano_ref = data_ref[:4]

        # Mapeamento col → mes do ano corrente
        meses_header = None
        commodity_atual = None

        rows_data = list(ws.iter_rows(values_only=True))
        for row_idx, row in enumerate(rows_data):
            # Pega celula textual mais a esquerda nao-vazia
            primeira_texto = None
            primeira_col = None
            for c_idx, c in enumerate(row[:5]):
                if c is not None and isinstance(c, str) and c.strip():
                    primeira_texto = c.strip()
                    primeira_col = c_idx
                    break

            if primeira_texto is None:
                continue

            primeira_lower = primeira_texto.lower()

            # Header de meses (linha onde aparecem "Fev", "Mar", "Abr"...)
            if any(m in primeira_lower or any(m in str(c).lower() if isinstance(c, str) else False for c in row[:13]) for m in ["fev", "mar", "abr", "mai"]):
                # Detectar quais cols sao meses
                col_meses = {}
                meses_abrev = {"jan": "01", "fev": "02", "mar": "03", "abr": "04",
                               "mai": "05", "jun": "06", "jul": "07", "ago": "08",
                               "set": "09", "out": "10", "nov": "11", "dez": "12"}
                for c_idx, c in enumerate(row):
                    if isinstance(c, str) and c.strip().lower()[:3] in meses_abrev:
                        col_meses[c_idx] = meses_abrev[c.strip().lower()[:3]]
                if col_meses and not meses_header:
                    meses_header = col_meses
                continue

            # Detectar mudanca de commodity (linhas tipo "1. Soja", "2. Farelo", "3. Oleo")
            m_secao = re.match(r"^\s*\d+\.\s*(soja|farelo|oleo|óleo|gr[aã]o)", primeira_lower)
            if m_secao:
                produto = m_secao.group(1).lower()
                commodity_atual = COMMODITIES_SLUG.get(produto, produto)
                continue

            if not commodity_atual or not meses_header:
                continue

            # Detectar metrica (linhas tipo "1.3. Processamento", "1.4 Exportacao")
            texto_clean = re.sub(r"^\s*\d+\.\d+\s*\.?\s*", "", primeira_texto).strip().lower()
            metrica = None
            for chave, slug in METRICAS_TARGET.items():
                if chave in texto_clean:
                    metrica = slug
                    break

            if not metrica:
                continue

            # Iterar pelos meses na linha
            for col_idx, mes_str in meses_header.items():
                if col_idx >= len(row):
                    continue
                valor = row[col_idx]
                if isinstance(valor, (int, float)) and valor != 0:
                    data_mes = f"{ano_ref}-{mes_str}-01"
                    results.append({
                        "data_referencia": data_mes,
                        "tipo": "supply_demand",
                        "commodity": f"{commodity_atual}_brasil",
                        "metrica": metrica,
                        "valor": float(valor),
                        "unidade": "mil_t",
                        "contexto": f"ABIOVE projecoes_mensais | {commodity_atual} | mes {mes_str}/{ano_ref}",
                    })

        return results

    def _erro(self, msg: str) -> dict:
        return {
            "data_referencia": date.today().isoformat(),
            "tipo": "erro",
            "commodity": "abiove",
            "metrica": "fetch_error",
            "valor": None,
            "contexto": msg,
        }
